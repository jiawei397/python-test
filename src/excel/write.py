import openpyxl
import os
import re
import logging

FORMAT_DATA = format = "%(asctime)s - %(name)s -[%(lineno)d] - %(message)s"
DEFINE_TIME = "%Y/%m/%d %H:%M:%S"  # 自定义时间格式
logging.basicConfig(level=logging.ERROR,
                    datefmt=DEFINE_TIME,
                    format=FORMAT_DATA,
                    filename="write.log",
                    filemode="a+")

file = 'demo.xlsx'
qingdan_dict = {}
end_row_map = {}
pattern = re.compile(r'^[-+]?[-0-9]\d*\.\d*|[-+]?\.?[0-9]\d*$')
map = {}


# 遍历指定目录，显示目录下的所有文件名
def eachFile(filepath):
    pathDir = os.listdir(filepath)
    num = 1
    for allDir in pathDir:
        child = os.path.join('%s/%s' % (filepath, allDir))
        # print(child)  # .decode('gbk')是解决中文显示乱码问题
        read_excel(child, num)
        num += 1


def is_number(num):
    result = pattern.match(str(num))
    if result:
        return True
    else:
        return False


def get_key(table, row, num):
    return table + '-' + str(row) + '-' + str(num)


def is_key(table, row, num):
    return get_key(table, row, num) in map


def read_excel(file, num):
    data = openpyxl.load_workbook(file)
    # print(data.get_named_ranges())  # 输出工作页索引范围
    # print(data.get_sheet_names())  # 输出所有工作页的名称
    # 取第一张表
    # sheetnames = data.get_sheet_names()

    # table = data.get_sheet_by_name('A3_01_SBL_PRC') #sheetnames[2]
    # print(data.worksheets)

    for table in data.worksheets:
        if table.title.find('清单') == -1 and table.title.find('填写须知') == -1:
            read_table(table)
        # elif table.title.find('清单') != -1:
        #     read_qingdan_table(table, num)

            # table = data.worksheets[2]

            # read_table(table)
            # values = ['E', 'X', 'C', 'E', 'L']
            # for value in value:
            #     table.cell(nrows + 1, 1).value = value
            #     nrows = nrows + 1
            # print(table.cell(4, 4).value)
            # table.cell(4, 4).value = 33
            # data.save('excel_test.xlsx')

            # list.append(table.cell(4, 4).value)


def read_table(table):
    # print(table.title)  # 输出表名
    nrows = table.max_row  # 获得行数
    ncolumns = table.max_column  # 获得行数

    for row in range(1, nrows):
        for col in range(1, ncolumns):
            val = table.cell(row, col).value
            if val and is_number(val):
                key = get_key(table.title, row, col)
                map.setdefault(key, [])
                map[key].append(val)


def write_excel(file):
    data = openpyxl.load_workbook(file)
    for table in data.worksheets:
        if table.title.find('清单') == -1 and table.title.find('填写须知') == -1:
            nrows = table.max_row  # 获得行数
            ncolumns = table.max_column  # 获得行数

            for row in range(1, nrows):
                for col in range(1, ncolumns):
                    if is_key(table.title, row, col):
                        try:
                            table.cell(row, col).value = sum(map[get_key(table.title, row, col)])
                        except AttributeError as e:
                            print(e, table.title, row, col)
                            logging.error(e)
                            logging.error('table=' + table.title + ' row=' + str(row) + ' col=' + str(col))

                            # TODO elif table.title.find('清单') != -1:
        # elif table.title == '4.应收清单20190930':
        #     #     print(table.title)
        #     for key in qingdan_dict:
        #         arr = key.split('-')
        #         if arr[0] == table.title:
        #             # print(key, qingdan_dict[key])
        #             table.cell(int(arr[1]), int(arr[2])).value = qingdan_dict[key]

    data.save('汇总表.xlsx')


# 读取清单
def read_qingdan_table(table, num):
    nrows = table.max_row  # 获得行数
    ncolumns = table.max_column  # 获得行数
    # print(nrows)

    end_row = nrows

    if table.title not in end_row_map:
        for row in range(3, nrows):
            for col in range(1, ncolumns):
                val = table.cell(row, col).value
                if val != None:
                    if val == '合计':
                        # print(row, col)
                        end_row = row - 1
                        end_row_map[table.title] = end_row
    else:
        end_row = end_row_map[table.title]

    for row in range(3, end_row):
        for col in range(1, ncolumns):
            val = table.cell(row, col).value
            if val != None:
                # print(row, col, val)
                key = get_key(table.title, row - 2 + (num - 1) * end_row, col)
                qingdan_dict[key] = val


# def read_qingdan(file):
#     data = openpyxl.load_workbook(file)
#     table = data.get_sheet_by_name('4.应收清单20190930')  # sheetnames[2]
#     read_qingdan_table(table)
#     print(qingdan_dict)

# read_qingdan(file)

eachFile('data')
write_excel(file)
# print(map)
